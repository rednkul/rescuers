import sys, os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

import datetime




from workers.models import Service, Worker, Division
from workers.views import State
from workers.templatetags.is_zero import is_zero

today = datetime.date.today()
month_names = ['января', 'февраля', 'марта', 'апреля', 'мая', 'июня', 'июля', 'августа', 'сентября', 'октября',
                   'ноября', 'декабря']
excel_file = openpyxl.load_workbook('media/excel/список.xlsx')
excel_sheet = excel_file['Лист1']

# Шрифты
divisions_font = Font(name='TimesNewRoman',
                    size=12,
                    bold=True,
                    italic=True,
                    )

table_font = Font(name='TimesNewRoman',
                    size=12,
                    )

bold_font = Font(name='TimesNewRoman',
                    size=12,
                    bold=True,
                    )
vacancy_font = Font(name='TimesNewRoman',
                    size=12,
                    bold=True,
                    color='ed0e0e',
                    )

# Границы
empty_fill = PatternFill()
empty_border = Border()

table_border = table_border =Border(left=Side(border_style='thin',
                          color='000000'),
                right=Side(border_style='thin',
                           color='000000'),
                top=Side(border_style='thin',
                         color='000000'),
                bottom=Side(border_style='thin',
                            color='000000'),
                              )

# Выравнивания
center_alignment = Alignment(horizontal='center', vertical='center')
right_alignment = Alignment(horizontal='right', vertical='center')
left_alignment = Alignment(horizontal='left', vertical='center')

# Заливки
division_total_fill = PatternFill(fill_type='solid', start_color='d9d9d9')
female_fill = PatternFill(fill_type='solid', start_color='ffff00')
rescuer_fill = PatternFill(fill_type='solid', start_color='92d050')
operative_fill = PatternFill(fill_type='solid', start_color='ffff00')

# Стили


# Начало редактируемой части
row_number = 7
columns = ['A', 'B', 'C', 'D']
class ListImportToExcel(State):
    def __init__(self):
        self.row_number = 7
        self.columns = ['A', 'B', 'C', 'D']

    def make_excel(self):

        self.make_header()

        for division in Division.objects.all():

            # Заголовок - название подразделения
            excel_sheet[f'A{self.row_number}'] = f'{division.name}'
            excel_sheet[f'A{self.row_number}'].font = divisions_font
            excel_sheet[f'A{self.row_number}'].alignment = center_alignment
            excel_sheet.merge_cells(f'A{self.row_number}:D{self.row_number}')

            self.row_number += 1

            # Пустая таблица после заголовка
            self.empty_row()

            self.row_number += 1
            # Список сотрудников подразделения
            for index, worker in enumerate(division.division_workers.all().order_by('post__priority')):

                excel_sheet[f'A{self.row_number}'] = index + 1
                excel_sheet[f'A{self.row_number}'].alignment = center_alignment

                if worker.name == 'Вакансия':
                    excel_sheet[f'B{self.row_number}'] = worker.name.upper()
                    excel_sheet[f'B{self.row_number}'].font = vacancy_font
                else:
                    excel_sheet[f'B{self.row_number}'] = worker.get_full_name()

                if worker.sex == 'ЖЕН':
                    excel_sheet[f'B{self.row_number}'].fill = female_fill
                excel_sheet[f'B{self.row_number}'].alignment = left_alignment

                excel_sheet[f'C{self.row_number}'] = worker.post.name
                excel_sheet[f'C{self.row_number}'].alignment = center_alignment
                if worker.post.rescuer:
                    excel_sheet[f'C{self.row_number}'].fill = rescuer_fill
                elif not worker.post.rescuer and worker.post.operative:
                    excel_sheet[f'C{self.row_number}'].fill = operative_fill

                if worker.date_beginning:
                    excel_sheet[f'D{self.row_number}'] = f'{"0"+str(worker.date_beginning.day) if worker.date_beginning.day < 10 else worker.date_beginning.day }.{worker.date_beginning.month}.{worker.date_beginning.year}'
                else:
                    excel_sheet[f'D{self.row_number}'] = None
                excel_sheet[f'D{self.row_number}'].alignment = center_alignment

                for column in self.columns:
                    excel_sheet[f'{column}{self.row_number}'].border = table_border
                    excel_sheet[f'{column}{self.row_number}'].font = table_font

                if worker.name == 'Вакансия':
                    excel_sheet[f'B{self.row_number}'] = worker.name.upper()
                    excel_sheet[f'B{self.row_number}'].font = vacancy_font
                    excel_sheet[f'C{self.row_number}'].font = vacancy_font
                else:
                    excel_sheet[f'B{self.row_number}'] = worker.get_full_name()

                excel_sheet[f'A{self.row_number}'].font = bold_font

                self.row_number += 1

            excel_sheet[f'A{self.row_number}'] = None
            excel_sheet[f'B{self.row_number}'] = None

            excel_sheet[f'C{self.row_number}'] = f'Шт. {division.get_state()}'


            excel_sheet[f'D{self.row_number}'] = f'Факт {Worker.objects.filter(division=division, on_duty=True).count()}'

            for column in self.columns[2:]:
                excel_sheet[f'{column}{self.row_number}'].fill = division_total_fill
                excel_sheet[f'{column}{self.row_number}'].alignment = left_alignment

            for column in self.columns:
                excel_sheet[f'{column}{self.row_number}'].border = table_border
                excel_sheet[f'{column}{self.row_number}'].font = bold_font

            self.row_number += 1

            self.empty_row()

            self.row_number += 1

            excel_sheet[f'D{self.row_number}'] = f'Списочно {Worker.objects.filter(division=division).exclude(name="Вакансия").count()}'

            excel_sheet[f'D{self.row_number}'].font = bold_font
            excel_sheet[f'D{self.row_number}'].alignment = left_alignment

            self.row_number += 1

            self.empty_row()

            self.row_number += 1


        self.make_total()
        excel_file.save(filename=f'media/excel/list_reports/Список работников филиала ЯВГСО на {today.day} {month_names[today.month - 1]} {today.year} года.xlsx')

    def empty_row(self):
        for column in self.columns:
            excel_sheet[f'{column}{self.row_number}'] = None
            excel_sheet[f'{column}{self.row_number}'].border = empty_border
            excel_sheet[f'{column}{self.row_number}'].fill = empty_fill


    def make_header(self):
        excel_sheet['A3'] = f'   на {today.day} {month_names[today.month - 1]} {today.year} года			'
        excel_sheet['B4'] = f'Женщин фактически - {Worker.objects.filter(sex="ЖЕН", on_duty=True).count()}'
        excel_sheet['B5'] = f'Женщин списочно - {Worker.objects.filter(sex="ЖЕН").count()}'
        excel_sheet['C4'] = f'Шт. {self.get_state_workers_number()}'
        excel_sheet['D4'] = f'Факт. {self.get_workers_number() - Worker.objects.filter(on_duty=False).count()}'
        excel_sheet['D5'] = f'Списочно {self.get_workers_number()}'
        excel_sheet['D6'] = f'Вакансий - {self.get_vacancies_number()}'


    def make_total(self):
        row_number = 8
        columns = ['F', 'G']
        excel_sheet[f'F{row_number}'] = f'Шт. {self.get_state_workers_number()}'
        excel_sheet[f'G{row_number}'] = f'Факт {self.get_workers_number()}'
        excel_sheet.column_dimensions['F'].width = 35
        excel_sheet.column_dimensions['G'].width = 13

        for column in columns:
            excel_sheet[f'{column}{row_number}'].font = bold_font
            excel_sheet[f'{column}{row_number}'].alignment = center_alignment
            for row in range(row_number + 1, row_number + 16):
                excel_sheet[f'{column}{row}'].font = table_font
                excel_sheet[f'{column}{row}'].alignment = center_alignment



        row_number += 1

        excel_sheet[f'F{row_number}'] = 'командир отряда'
        excel_sheet[f'G{row_number}'] = Worker.objects.filter(post__name__iregex="Командир отряда").exclude(name='Вакансия').count()

        row_number += 1

        excel_sheet[f'F{row_number}'] = 'заместитель командира отряда'
        excel_sheet[f'G{row_number}'] = Worker.objects.filter(post__name__iregex="Заместитель командира отряда").count()

        row_number += 1
        self.empty_row()
        row_number += 1

        excel_sheet[f'F{row_number}'] = 'помощник командира отряда'
        excel_sheet[f'G{row_number}'] = Worker.objects.filter(post__name__iregex="Помощник командира отряда").count()

        row_number += 1
        self.empty_row()
        row_number += 1

        excel_sheet[f'F{row_number}'] = 'заместитель командира взвода'
        excel_sheet[f'G{row_number}'] = Worker.objects.filter(post__name__iregex="Заместитель командира взвода").count()

        row_number += 1

        excel_sheet[f'F{row_number}'] = 'помощник командира взвода'
        excel_sheet[f'G{row_number}'] = Worker.objects.filter(post__name__iregex="помощник командира взвода").count()

        row_number += 1

        excel_sheet[f'F{row_number}'] = 'водитель автомобиля (специального)'
        excel_sheet[f'F{row_number}'].alignment = Alignment(horizontal='center', wrap_text=True)
        excel_sheet[f'G{row_number}'] = Worker.objects.filter(post__name__iregex="Водитель автомобиля (специального)").count()

        row_number += 1

        excel_sheet[f'F{row_number}'] = 'командир отделения'
        excel_sheet[f'G{row_number}'] = Worker.objects.filter(post__name__iregex="Командир отделения").exclude(name='Вакансия').count()

        row_number += 1

        excel_sheet[f'F{row_number}'] = 'респираторщик'
        excel_sheet[f'G{row_number}'] = Worker.objects.filter(post__name__iregex="Респираторщик").exclude(name='Вакансия').count()

        row_number += 1

        excel_sheet[f'F{row_number}'] = 'Всего по штату'
        excel_sheet[f'F{row_number}'].font = bold_font
        excel_sheet[f'G{row_number}'] = Worker.objects.filter(post__operative=True).count()

        for row in range(row_number, row_number + 6):
            excel_sheet[f'F{row}'].alignment = left_alignment

        row_number += 1

        excel_sheet[f'F{row_number}'] = 'Вакансия респираторщик'
        excel_sheet[f'G{row_number}'] = Worker.objects.filter(name='Вакансия', post__name='Респираторщик').count()

        row_number += 1

        excel_sheet[f'F{row_number}'] = 'Всего списочно'
        excel_sheet[f'F{row_number}'].font = bold_font
        excel_sheet[f'G{row_number}'] = self.get_state_operative_number()

        row_number += 1

        excel_sheet[f'F{row_number}'] = 'Респираторный состав штат'
        excel_sheet[f'G{row_number}'] = self.get_resperator_state()

        row_number += 1

        excel_sheet[f'F{row_number}'] = 'Респираторный состав список'
        excel_sheet[f'G{row_number}'] = self.get_resperator_number()




file = ListImportToExcel()
file.make_excel()