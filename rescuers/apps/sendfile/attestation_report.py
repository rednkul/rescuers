import sys, os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

import datetime




from workers.models import Division, Worker
from workers.views import State
from workers.templatetags.is_zero import is_zero
from workers.templatetags.time_to_attestation import next_attestation_excel

today = datetime.date.today()

excel_file = openpyxl.load_workbook('media/excel/Аттестация.xlsx')
excel_sheet = excel_file['Лист1']

# Шрифты
divisions_font = Font(name='TimesNewRoman',
                    size=12,
                    bold=True,
                    italic=True,
                    )

table_font = Font(name='TimesNewRoman',
                    size=12,
                    )

bold_font = Font(name='TimesNewRoman',
                    size=12,
                    bold=True,
                    )
division_head_font = Font(name='TimeNewRoman',
                          size=12,

                          bold=True,
                          italic=True)
division_table_fill = PatternFill(fill_type='solid', start_color='176e25')

division_head_fill = PatternFill(fill_type='solid', start_color='7875ff')

# Границы
empty_fill = PatternFill()
empty_border = Border()

table_border  =Border(left=Side(border_style='thin',
                          color='000000'),
                right=Side(border_style='thin',
                           color='000000'),
                top=Side(border_style='thin',
                         color='000000'),
                bottom=Side(border_style='thin',
                            color='000000'),
                              )


# Выравнивания
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
right_alignment = Alignment(horizontal='right', vertical='center',)
left_alignment = Alignment(horizontal='left', vertical='center')

# Заливки
division_total_fill = PatternFill(fill_type='solid', start_color='d9d9d9')
female_fill = PatternFill(fill_type='solid', start_color='ffff00')
rescuer_fill = PatternFill(fill_type='solid', start_color='92d050')
operative_fill = PatternFill(fill_type='solid', start_color='ffff00')

# Стили


# Начало редактируемой части


soon_attestation_filter = datetime.datetime.today() + datetime.timedelta(days=90) - datetime.timedelta(
                days=365 * 3)

class AttestationImportToExcel(State):
    def __init__(self):
        self.row_number = 1
        self.columns = ['A', 'B', 'C', 'D', 'E']

    def make_excel(self):

        self.make_header()
        self.row_number += 2
        for division in Division.objects.all():

            # Заголовок - название подразделения
            excel_sheet[f'A{self.row_number}'] = f'{division.name} '
            excel_sheet[f'A{self.row_number}'].font = divisions_font
            excel_sheet[f'A{self.row_number}'].alignment = center_alignment
            excel_sheet.merge_cells(f'A{self.row_number}:E{self.row_number}')
            for column in self.columns:
                excel_sheet[f'{column}{self.row_number}'].border = table_border
                excel_sheet[f'{column}{self.row_number}'].font = bold_font
                excel_sheet[f'{column}{self.row_number}'].fill = division_head_fill
            self.row_number += 1

            excel_sheet.row_dimensions[self.row_number].height = 35
            excel_sheet[f'A{self.row_number}'] = '№ п/п'
            excel_sheet[f'B{self.row_number}'] = 'Сотрудник'
            excel_sheet[f'C{self.row_number}'] = 'Должность'
            excel_sheet[f'D{self.row_number}'] = 'Дата последней аттестации'
            excel_sheet[f'E{self.row_number}'] = 'Дата следующей аттестации'

            for column in self.columns:

                excel_sheet[f'{column}{self.row_number}'].border = table_border
                excel_sheet[f'{column}{self.row_number}'].font = division_head_font
                excel_sheet[f'{column}{self.row_number}'].fill = division_table_fill
                excel_sheet[f'{column}{self.row_number}'].alignment = center_alignment
            self.row_number += 1
            # Список сотрудников подразделения
            for index, worker in enumerate(division.division_workers.filter(date_attestation__lt=soon_attestation_filter).order_by('post__priority')):

                excel_sheet[f'A{self.row_number}'] = index + 1


                excel_sheet[f'B{self.row_number}'] = worker.get_full_name()



                excel_sheet[f'C{self.row_number}'] = worker.post.name



                excel_sheet[f'D{self.row_number}'] = f'{"0"+str(worker.date_beginning.day) if worker.date_beginning.day < 10 else worker.date_beginning.day }.{worker.date_beginning.month}.{worker.date_beginning.year}'
                excel_sheet[f'D{self.row_number}'].alignment = center_alignment

                excel_sheet[f'E{self.row_number}'] = next_attestation_excel(worker.date_attestation)


                for column in self.columns:
                    excel_sheet[f'{column}{self.row_number}'].border = table_border
                    excel_sheet[f'{column}{self.row_number}'].font = table_font
                    excel_sheet[f'{column}{self.row_number}'].alignment = center_alignment

                excel_sheet[f'A{self.row_number}'].font = bold_font

                self.row_number += 1



            excel_sheet[f'D{self.row_number}'] = f'Всего в подразделении'
            excel_sheet[f'E{self.row_number}'] = division.division_workers.filter(date_attestation__lt=soon_attestation_filter).count()
            for column in self.columns[3:]:
                excel_sheet[f'{column}{self.row_number}'].border = table_border
                excel_sheet[f'{column}{self.row_number}'].font = table_font
                excel_sheet[f'{column}{self.row_number}'].alignment = center_alignment
                excel_sheet[f'{column}{self.row_number}'].fill = division_total_fill
            self.row_number += 1



            self.row_number += 1


        self.make_total()
        excel_file.save(filename=f'media/excel/attestation_reports/Аттестующиеся {today.day}.{today.month}.{today.year}.xlsx')

    def empty_row(self):
        for column in self.columns:
            excel_sheet[f'{column}{self.row_number}'] = None
            excel_sheet[f'{column}{self.row_number}'].border = empty_border
            excel_sheet[f'{column}{self.row_number}'].fill = empty_fill


    def make_header(self):
        excel_sheet['A1'] = f'Список работников ЯВГСО, аттестующихся в ближайшие 90 дней   на {today.day}.{today.month}.{today.year} года			'
        excel_sheet.merge_cells(f'A1:E1')
        for column in self.columns:
            excel_sheet[f'{column}{self.row_number}'].border = table_border
            excel_sheet[f'{column}{self.row_number}'].fill = division_head_fill


    def make_total(self):


        excel_sheet['H1'] = 'Всего по отряду:'
        excel_sheet['I1'] = Worker.objects.filter(date_attestation__lt=soon_attestation_filter).exclude(name="Вакансия").count()

        for column in ['H', 'I']:
            excel_sheet[f'{column}1'].border = table_border
            excel_sheet[f'{column}1'].font = bold_font
            excel_sheet[f'{column}1'].alignment = center_alignment
            excel_sheet[f'{column}1'].fill = division_head_fill







