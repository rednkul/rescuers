import sys, os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

import datetime


from workers.models import Service, Worker
from workers.views import State
from workers.templatetags.is_zero import is_zero


excel_file = openpyxl.load_workbook('media/excel/Укомплектованность.xlsx')

excel_sheet = excel_file['Укомплектованность']


excel_sheet['C4'] = f'УКОМПЛЕКТОВАННОСТЬ филиала                                   "Якутский ВГСО" ФГУП "ВГСЧ" на {datetime.datetime.now().year}.{datetime.datetime.now().month}.{datetime.datetime.now().day}'


row_number = 6

cells_font = Font(name='TimesNewRoman',
                    size=12,
                    )
headers_font = Font(name='TimesNewRoman',
                    size=12,
                    bold=True,
                    )

headers_alignment = Alignment(horizontal='center',)

totals_alignment = Alignment(horizontal='right',)
totals_fill = PatternFill(fill_type='solid',start_color='FFFF00')

admin_fill = PatternFill(fill_type='solid', start_color='fce4d6')
columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

explanation_border = Border(left=Side(border_style='thin',
                          color='ffffff'),
                right=Side(border_style='thin',
                           color='ffffff'),
                top=Side(border_style='thin',
                         color='ffffff'),
                bottom=Side(border_style='thin',
                            color='ffffff'),
                              )

table_border =Border(left=Side(border_style='thin',
                          color='000000'),
                right=Side(border_style='thin',
                           color='000000'),
                top=Side(border_style='thin',
                         color='000000'),
                bottom=Side(border_style='thin',
                            color='000000'),
                              )

numbers_alignment = Alignment(horizontal='center', )

class ImportToExcel(State):
    def __init__(self):
        self.excel_file = openpyxl.load_workbook('media/excel/Укомплектованность.xlsx')

        self.excel_sheet = excel_file['Укомплектованность']

        self.excel_sheet[
            'C4'] = f'УКОМПЛЕКТОВАННОСТЬ филиала                                   "Якутский ВГСО" ФГУП "ВГСЧ" на {datetime.datetime.now().strftime("%Y.%m.%d")}'
        self.excel_sheet['E3'] = f' от______________{datetime.datetime.now().year} №______'

        self.row_number = 6

        self.headers_font = Font(name='TimesNewRoman',
                            size=12,
                            bold=True,
                            )

        self.headers_alignment = Alignment(horizontal='center', )

        self.explanation_alignment = Alignment(horizontal='left', wrap_text=True)

        self.totals_alignment = Alignment(horizontal='right', )
        self.totals_fill = PatternFill(fill_type='solid', start_color='FFFF00')

        self.admin_fill = PatternFill(fill_type='solid', start_color='fce4d6')
        self.columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    def make_excel(self):
        for i in range(Service.objects.all().count()):
            for column in self.columns:
                excel_sheet[f'{column}{self.row_number}'].font = self.headers_font
                excel_sheet[f'{column}{self.row_number}'].alignment = self.headers_alignment
                excel_sheet[f'{column}{self.row_number}'].border = table_border
            excel_sheet[f'A{self.row_number}'] = f'{i + 1}'
            excel_sheet[f'B{self.row_number}'] = ''
            excel_sheet[f'C{self.row_number}'] = Service.objects.all()[i].name
            excel_sheet[f'D{self.row_number}'] = ''
            excel_sheet[f'E{self.row_number}'] = ''
            excel_sheet[f'F{self.row_number}'] = ''
            excel_sheet[f'G{self.row_number}'] = ''
            excel_sheet[f'H{self.row_number}'] = ''
            self.row_number += 1
            for j in range(Service.objects.all()[i].service_posts.all().count()):
                excel_sheet[f'A{self.row_number}'] = f'{i + 1}.{j + 1}'
                if Service.objects.all()[i].service_posts.all()[j].operative:
                    excel_sheet[f'B{self.row_number}'] = 'оперативный'
                else:
                    excel_sheet[f'B{self.row_number}'] ='административно-технический'
                    excel_sheet[f'B{self.row_number}'].fill = self.admin_fill
                excel_sheet[f'C{self.row_number}'] = Service.objects.all()[i].service_posts.all()[j].name
                excel_sheet[f'D{self.row_number}'] = is_zero(Service.objects.all()[i].service_posts.all()[j].get_post_state())
                excel_sheet[f'E{self.row_number}'] = is_zero(Service.objects.all()[i].service_posts.all()[j].get_post_workers_number())
                excel_sheet[f'F{self.row_number}'] = is_zero(Service.objects.all()[i].service_posts.all()[j].get_post_vacancies_number())
                excel_sheet[f'G{self.row_number}'] = is_zero(Service.objects.all()[i].service_posts.all()[j].get_overstaffing())
                excel_sheet[f'H{self.row_number}'] = is_zero(Service.objects.all()[i].service_posts.all()[j].get_staffing_percent())
                for column in columns:
                    excel_sheet[f'{column}{self.row_number}'].border = table_border
                    excel_sheet[f'{column}{self.row_number}'].font = cells_font
                excel_sheet[f'A{self.row_number}'].alignment = numbers_alignment
                self.row_number += 1
            excel_sheet[f'A{self.row_number}'] = ''
            excel_sheet[f'B{self.row_number}'] = ''

            excel_sheet[f'C{self.row_number}'] = f'Итого по {Service.objects.all()[i].get_abbreviation()}:'
            excel_sheet[f'C{self.row_number}'].alignment = self.totals_alignment
            excel_sheet[f'C{self.row_number}'].font = self.headers_font
            for column in columns:
                excel_sheet[f'{column}{self.row_number}'].border = table_border


            for column in columns[3:]:
                excel_sheet[f'{column}{self.row_number}'].fill = self.totals_fill
                excel_sheet[f'{column}{self.row_number}'].font = cells_font

            excel_sheet[f'D{self.row_number}'] = Service.objects.all()[i].get_service_state()
            excel_sheet[f'E{self.row_number}'] = Service.objects.all()[i].get_service_workers_number()
            excel_sheet[f'F{self.row_number}'] = Service.objects.all()[i].get_service_vacancies_number()
            excel_sheet[f'G{self.row_number}'] = Service.objects.all()[i].get_service_overstaffing()
            excel_sheet[f'H{self.row_number}'] = Service.objects.all()[i].get_service_staffing_percent()
            self.row_number += 1

        # Общее по штату
        excel_sheet[f'C{self.row_number}'] = 'ВСЕГО:'
        excel_sheet[f'D{self.row_number}'] = self.get_state_workers_number()
        excel_sheet[f'E{self.row_number}'] = self.get_workers_number()
        excel_sheet[f'F{self.row_number}'] = self.get_vacancies_number()
        excel_sheet[f'G{self.row_number}'] = self.get_overstaffing()
        excel_sheet[f'H{self.row_number}'] = self.get_staffing_percent()

        excel_sheet[f'C{self.row_number}'].font = self.headers_font
        excel_sheet[f'C{self.row_number}'].alignment = self.totals_alignment

        for column in columns[3:]:
            excel_sheet[f'{column}{self.row_number}'].font = cells_font

        for column in columns:
            excel_sheet[f'{column}{self.row_number}'].border = table_border

        self.row_number += 1

        # По оперативному составу
        excel_sheet[f'B{self.row_number}'] = 'оперативный'
        excel_sheet[f'B{self.row_number}'].font = cells_font
        excel_sheet[f'C{self.row_number}'] = 'в том числе по оперативному составу:'
        excel_sheet[f'D{self.row_number}'] = self.get_state_operative_number()
        excel_sheet[f'E{self.row_number}'] = self.get_operative_number()
        excel_sheet[f'F{self.row_number}'] = self.get_operative_vacancies_number()
        excel_sheet[f'G{self.row_number}'] = self.get_operative_overstaffing_number()
        excel_sheet[f'H{self.row_number}'] = self.get_operative_staffing_percent()

        excel_sheet[f'C{self.row_number}'].font = self.headers_font
        excel_sheet[f'C{self.row_number}'].alignment = self.totals_alignment

        for column in columns[3:]:
            excel_sheet[f'{column}{self.row_number}'].font = cells_font

        for column in columns:
            excel_sheet[f'{column}{self.row_number}'].border = table_border

        self.row_number += 1

        # По Ад-Тех составу
        excel_sheet[f'B{self.row_number}'] = 'административно-технический'
        excel_sheet[f'B{self.row_number}'].font = cells_font
        excel_sheet[f'B{self.row_number}'].fill = self.admin_fill
        excel_sheet[f'C{self.row_number}'] = 'в том числе по административно-техническому составу:'
        excel_sheet[f'D{self.row_number}'] = self.get_state_admin_number()
        excel_sheet[f'E{self.row_number}'] = self.get_admin_number()
        excel_sheet[f'F{self.row_number}'] = self.get_admin_vacancies_number()
        excel_sheet[f'G{self.row_number}'] = self.get_admin_overstaffing_number()
        excel_sheet[f'H{self.row_number}'] = self.get_admin_staffing_percent()

        excel_sheet[f'C{self.row_number}'].font = self.headers_font
        excel_sheet[f'C{self.row_number}'].alignment = self.totals_alignment

        for column in columns[3:]:
            excel_sheet[f'{column}{self.row_number}'].font = cells_font

        for column in columns:
            excel_sheet[f'{column}{self.row_number}'].border = table_border

        self.row_number += 1

        for row in range(7):
            for column in columns:
                excel_sheet[f'{column}{row+self.row_number}'].border = explanation_border
                excel_sheet[f'{column}{row+self.row_number}'].font = cells_font

        excel_sheet.merge_cells(f'A{self.row_number}:C{self.row_number}')
        excel_sheet[f'A{self.row_number}'] = '* - работники Учебного центра, непривлекаемые к горноспасательным работам на обслуживаемых предприятиях (неаттестованные на право ведения аварийно-спасательных работ(горноспасательных работ) '
        excel_sheet.row_dimensions[self.row_number].height = 48
        excel_sheet[f'A{self.row_number}'].alignment = self.explanation_alignment

        self.row_number += 1

        for column in self.columns:
            excel_sheet[f'{column}{self.row_number}'] = ''

        self.row_number += 1

        if Worker.objects.filter(post__name="Командир отряда"):
            commander = Worker.objects.filter(post__name="Командир отряда")[0]
            excel_sheet[f'A{self.row_number}'] = f'Командир отряда филиала "Якутский ВГСО" ФГУП "ВГСЧ"                                {commander.get_initials()}'

        excel_sheet.merge_cells(f'A{self.row_number}:H{self.row_number}')
        excel_sheet[f'A{self.row_number}'].alignment = self.explanation_alignment

        self.row_number += 1

        for column in self.columns:
            excel_sheet[f'{column}{self.row_number}'] = ''


        self.row_number += 1

        if Worker.objects.filter(post__name="Старший инспектор по кадрам"):
            inspector = Worker.objects.filter(post__name="Старший инспектор по кадрам")[0]
            excel_sheet[f'A{self.row_number}'] = f'Командир отряда филиала "Якутский ВГСО" ФГУП "ВГСЧ"                                {inspector.get_initials()}'


        excel_sheet.merge_cells(f'A{self.row_number}:H{self.row_number}')
        excel_sheet[f'A{self.row_number}'].alignment = self.explanation_alignment


        excel_file.save(filename=f'media/excel/staffing_reports/Укомплектованность на {datetime.datetime.now().strftime("%d.%m.%Y")}.xlsx')


file = ImportToExcel()
file.make_excel()